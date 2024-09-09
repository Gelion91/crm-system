import os

import requests
from bs4 import BeautifulSoup
from django.utils import dateformat
from django.core.files.base import ContentFile, File
from openpyxl import load_workbook
from core.models import Course, Logistics, Invoices
from crm import settings
from crm.settings import MEDIA_ROOT, BASE_DIR


def get_course():
    try:
        url = 'https://ligovka.ru/detailed/usd'
        response = requests.get(url)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        td = soup.find_all(class_='money_price')
        crs = Course(course=td[1].text)
        crs.save()
    except Exception as error:
        print(error)
    return Course.objects.last()


def get_invoice(logistic):
    logistic = Logistics.objects.get(pk=logistic.pk)
    wb = load_workbook(filename='other_files/Накладная_форма.xlsx')
    new_wb = wb
    ws = new_wb.active
    try:
        date_invoice = logistic.marker.split('-')
        date_invoice = date_invoice[-1] if len(date_invoice[-1]) >= 4 else date_invoice[-2]
        date_invoice = f"{date_invoice[2::]}-{date_invoice[:2]}-2024"
    except:
        date_invoice = 'Нет даты'

    ws['A7'] = date_invoice
    ws['B7'] = logistic.marker
    ws['D7'] = logistic.places
    ws['E7'] = logistic.weight
    ws['F7'] = logistic.tariff_one_kg
    ws['G7'] = logistic.tariff
    ws['I7'] = logistic.insurance
    ws['K7'] = logistic.package_price
    ws['M7'] = logistic.full_price
    new_wb.save(filename=f'other_files/{logistic.marker}.xlsx')
    with open(f'other_files/{logistic.marker}.xlsx', 'rb') as f:
        if Invoices.objects.filter(logistic=logistic).first():
            invoice = Invoices.objects.filter(logistic=logistic).first()
            # os.remove(path=str(BASE_DIR) + invoice.file.url)
        else:
            invoice = Invoices()
        invoice.logistic = logistic
        invoice.file = File(f, name=f'{logistic.marker}.xlsx')
        invoice.save()
        return invoice.file

